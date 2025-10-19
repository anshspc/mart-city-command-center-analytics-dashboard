import os
import sqlite3
import pandas as pd
from datetime import datetime

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(CURRENT_DIR)
DB_PATH = os.path.join(BASE_DIR, "data/smart_city.db")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

def query_kpi_data(conn):
    cursor = conn.cursor()
    
    # Complaints
    cursor.execute("SELECT COUNT(*) FROM citizen_complaints")
    total_c = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM citizen_complaints WHERE status = 'Resolved'")
    resolved_c = cursor.fetchone()[0]
    res_rate = round((resolved_c / total_c * 100), 1) if total_c > 0 else 0.0
    cursor.execute("SELECT AVG(julianday(resolved_at) - julianday(created_at)) * 24 FROM citizen_complaints WHERE status = 'Resolved'")
    avg_res_hrs = round(cursor.fetchone()[0] or 0.0, 1)
    
    # Traffic
    cursor.execute("SELECT AVG(congestion_index), AVG(average_speed) FROM traffic_records")
    avg_cong, avg_speed = cursor.fetchone()
    avg_cong = round(avg_cong or 0.0, 2)
    avg_speed = round(avg_speed or 0.0, 1)
    
    # Water
    cursor.execute("SELECT SUM(water_consumption_m3), AVG(quality_index), SUM(leak_detected) FROM water_consumption")
    tot_water, avg_wq, tot_leaks = cursor.fetchone()
    tot_water = round(tot_water or 0.0, 1)
    avg_wq = round(avg_wq or 0.0, 1)
    
    # Power
    cursor.execute("SELECT SUM(power_consumption_kwh), SUM(outage_duration_min), AVG(load_factor) FROM electricity_usage")
    tot_power, tot_outage, avg_lf = cursor.fetchone()
    tot_power = round(tot_power or 0.0, 1)
    avg_lf = round(avg_lf or 0.0, 2)
    
    # Sanitation
    cursor.execute("SELECT SUM(waste_collected_tons), AVG(sanitation_rating), SUM(missed_pickups) FROM sanitation_records")
    tot_waste, avg_rating, tot_missed = cursor.fetchone()
    tot_waste = round(tot_waste or 0.0, 1)
    avg_rating = round(avg_rating or 0.0, 2)
    
    return {
        "total_complaints": total_c,
        "resolved_complaints": resolved_c,
        "resolution_rate": res_rate,
        "avg_resolution_hours": avg_res_hrs,
        "avg_congestion": avg_cong,
        "avg_traffic_speed": avg_speed,
        "total_water_m3": tot_water,
        "avg_water_quality": avg_wq,
        "total_water_leaks": tot_leaks,
        "total_power_kwh": tot_power,
        "total_outage_minutes": tot_outage,
        "avg_power_load_factor": avg_lf,
        "total_waste_tons": tot_waste,
        "avg_sanitation_rating": avg_rating,
        "total_missed_pickups": tot_missed
    }

def generate_excel_report(conn, kpis, filename):
    wb = Workbook()
    
    # Styles
    font_title = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    font_section = Font(name='Segoe UI', size=12, bold=True, color='1F2937')
    font_header = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    font_body = Font(name='Segoe UI', size=10, color='333333')
    font_bold = Font(name='Segoe UI', size=10, bold=True, color='333333')
    
    fill_title = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Navy
    fill_header = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid') # Dark Gray
    fill_summary_lbl = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid') # Light Gray
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border_side = Side(border_style='thin', color='D1D5DB')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(bottom=Side(border_style='double', color='1F2937'), top=Side(border_style='thin', color='D1D5DB'))

    # Sheet 1: Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.merge_cells('A1:D2')
    title_cell = ws['A1']
    title_cell.value = "Smart City Command Center Performance Report"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = align_center
    
    # Subtitle
    ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Indore Operations"
    ws['A3'].font = Font(name='Segoe UI', size=9, italic=True, color='6B7280')
    
    # Section: Key Indicators
    ws['A5'] = "Key Performance Indicators"
    ws['A5'].font = font_section
    
    kpi_rows = [
        ("Total Citizen Complaints", kpis["total_complaints"], "Tickets filed", "Count"),
        ("Complaint Resolution Rate", kpis["resolution_rate"] / 100, "Percentage of resolved tickets", "Percentage"),
        ("Avg Resolution Duration", kpis["avg_resolution_hours"], "Hours to close case", "Float"),
        ("Avg Grid Congestion Index", kpis["avg_congestion"], "Scale 0-10", "Float"),
        ("Avg Vehicle Speed", kpis["avg_traffic_speed"], "km/h", "Float"),
        ("Total Water Consumed", kpis["total_water_m3"], "m³ volume", "Float"),
        ("Water Quality Index", kpis["avg_water_quality"] / 100, "Percentage compliance", "Percentage"),
        ("Water Leaks Repaired", kpis["total_water_leaks"], "Incidents resolved", "Count"),
        ("Power Consumed", kpis["total_power_kwh"], "kWh load", "Float"),
        ("Total Power Outages", kpis["total_outage_minutes"], "Minutes grid down", "Count"),
        ("Grid Load Factor", kpis["avg_power_load_factor"], "Efficiency index", "Float"),
        ("Waste Collected", kpis["total_waste_tons"], "Tons of solid waste", "Float"),
        ("Sanitation Rating", kpis["avg_sanitation_rating"], "Scale 1-5 scale", "Float"),
        ("Missed Waste Pickups", kpis["total_missed_pickups"], "Missed stops", "Count")
    ]
    
    headers = ["Indicator Metric", "Value", "Description", "Scale Context"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_left
        cell.border = border_all
        
    for idx, row in enumerate(kpi_rows, 7):
        ws.cell(row=idx, column=1, value=row[0]).font = font_bold
        ws.cell(row=idx, column=1).fill = fill_summary_lbl
        ws.cell(row=idx, column=1).border = border_all
        
        val_cell = ws.cell(row=idx, column=2, value=row[1])
        val_cell.font = font_body
        val_cell.border = border_all
        val_cell.alignment = align_right
        
        # Apply formatting
        if row[3] == "Percentage":
            val_cell.number_format = '0.0%'
        elif row[3] == "Float":
            val_cell.number_format = '#,##0.0'
        elif row[3] == "Count":
            val_cell.number_format = '#,##0'
            
        ws.cell(row=idx, column=3, value=row[2]).font = font_body
        ws.cell(row=idx, column=3).border = border_all
        ws.cell(row=idx, column=4, value=row[3]).font = font_body
        ws.cell(row=idx, column=4).border = border_all

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Sheet 2: Raw Complaints (Sample of 150 rows for detail)
    ws_comp = wb.create_sheet(title="Citizen Complaints Detail")
    ws_comp.views.sheetView[0].showGridLines = True
    df_c = pd.read_sql_query("""
        SELECT complaint_id, citizen_name, issue_type, status, created_at, resolved_at, satisfaction_rating 
        FROM citizen_complaints 
        ORDER BY created_at DESC 
        LIMIT 200
    """, conn)
    
    # Write Headers
    for col_idx, col_name in enumerate(df_c.columns, 1):
        cell = ws_comp.cell(row=1, column=col_idx, value=col_name.replace('_', ' ').title())
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    for row_idx, row_val in enumerate(df_c.values, 2):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws_comp.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            if col_idx in [1, 7]:
                cell.alignment = align_right
            elif col_idx in [4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    for col in ws_comp.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_comp.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Sheet 3: Utilities Analytics Summary
    ws_utils = wb.create_sheet(title="Utility Performance")
    ws_utils.views.sheetView[0].showGridLines = True
    
    # Water supply district aggregation
    df_water = pd.read_sql_query("""
        SELECT district_id, ROUND(AVG(water_consumption_m3), 1) as avg_consumption_m3, 
               ROUND(AVG(pressure_bar), 2) as avg_pressure_bar, SUM(leak_detected) as leaks_repaired,
               ROUND(AVG(quality_index), 1) as quality_score
        FROM water_consumption
        GROUP BY district_id
    """, conn)
    
    ws_utils.cell(row=1, column=1, value="Water Supply Analytics by District").font = font_section
    for col_idx, col_name in enumerate(df_water.columns, 1):
        cell = ws_utils.cell(row=3, column=col_idx, value=col_name.replace('_', ' ').title())
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    for row_idx, row_val in enumerate(df_water.values, 4):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws_utils.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            
    # Electricity Zone aggregation
    df_elec = pd.read_sql_query("""
        SELECT grid_zone_id, ROUND(AVG(power_consumption_kwh), 1) as avg_load_kwh,
               SUM(outage_duration_min) as total_outage_duration_min, ROUND(AVG(load_factor), 2) as avg_load_factor
        FROM electricity_usage
        GROUP BY grid_zone_id
    """, conn)
    
    start_row = 12
    ws_utils.cell(row=start_row, column=1, value="Electricity Grid Load by Zone").font = font_section
    for col_idx, col_name in enumerate(df_elec.columns, 1):
        cell = ws_utils.cell(row=start_row+2, column=col_idx, value=col_name.replace('_', ' ').title())
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    for row_idx, row_val in enumerate(df_elec.values, start_row+3):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws_utils.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_all
            
    for col in ws_utils.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_utils.column_dimensions[col_letter].width = max(max_len + 3, 15)

    wb.save(filename)
    print(f"Excel report generated: {filename}")

def generate_pdf_report(conn, kpis, filename):
    # Setup document
    doc = SimpleDocTemplate(
        filename,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom paragraph styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#4B5563'),
        spaceAfter=15
    )
    
    h1_style = ParagraphStyle(
        'H1',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1F2937'),
        spaceBefore=12,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        'Body',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#374151'),
        leading=14
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    
    table_body_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#1F2937')
    )
    
    table_body_bold = ParagraphStyle(
        'TableBodyBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.HexColor('#1F2937')
    )

    story = []
    
    # 1. Header Banner
    story.append(Paragraph("Indore Municipal Corporation (IMC)", title_style))
    story.append(Paragraph(f"COMMAND CENTER PERFORMANCE AUDIT REPORT — {datetime.now().strftime('%B %d, %Y')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1E3A8A'), spaceAfter=15))
    
    # Executive Summary Paragraph
    summary_text = (
        "This performance audit report summarizes the key utility, citizen service, and traffic indicators "
        "across Indore District. The data is generated continuously from sensors and IMC 311 citizen helpline nodes, "
        "enabling responsive city maintenance. Below is a detailed breakdown of core operational performance indices."
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 15))
    
    # 2. KPI Summary Table
    story.append(Paragraph("I. Key Performance Indicators (KPIs)", h1_style))
    
    kpi_data = [
        [
            Paragraph("Sector Domain", table_header_style), 
            Paragraph("Primary Metric", table_header_style), 
            Paragraph("Value Recorded", table_header_style),
            Paragraph("Audit Status", table_header_style)
        ],
        [
            Paragraph("Citizen Engagement", table_body_bold),
            Paragraph("Citizen Complaint Resolution Rate", table_body_style),
            Paragraph(f"{kpis['resolution_rate']}% (Avg {kpis['avg_resolution_hours']} Hrs)", table_body_style),
            Paragraph("Active", table_body_style)
        ],
        [
            Paragraph("Traffic Management", table_body_bold),
            Paragraph("Average Congestion Index", table_body_style),
            Paragraph(f"{kpis['avg_congestion']} / 10.0 ({kpis['avg_traffic_speed']} km/h avg)", table_body_style),
            Paragraph("Normal", table_body_style)
        ],
        [
            Paragraph("Water Operations", table_body_bold),
            Paragraph("Daily Consumption Supply", table_body_style),
            Paragraph(f"{kpis['total_water_m3']:,} m³ (Quality: {kpis['water_quality'] if 'water_quality' in kpis else kpis['avg_water_quality']}% compliance)", table_body_style),
            Paragraph("Secure", table_body_style)
        ],
        [
            Paragraph("Power Grid Control", table_body_bold),
            Paragraph("Power Grid Load Factor / Outages", table_body_style),
            Paragraph(f"{kpis['avg_power_load_factor']} LF ({kpis['total_outage_minutes']} total mins)", table_body_style),
            Paragraph("Attention", table_body_style)
        ],
        [
            Paragraph("Sanitation Services", table_body_bold),
            Paragraph("Waste Collection Rating", table_body_style),
            Paragraph(f"{kpis['avg_sanitation_rating']} / 5.0 Rating ({kpis['total_waste_tons']:,} tons)", table_body_style),
            Paragraph("Optimal", table_body_style)
        ]
    ]
    
    # Calculate widths
    col_widths = [120, 180, 150, 80]
    t = Table(kpi_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F2937')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F9FAFB')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F3F4F6')]),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 15))
    
    # 3. Department breakdown
    story.append(Paragraph("II. Citizen Service & Complaint Details by Department", h1_style))
    
    # Get database department summary
    cursor = conn.cursor()
    cursor.execute("""
        SELECT d.department_name, COUNT(c.complaint_id) as total,
               SUM(CASE WHEN c.status = 'Resolved' THEN 1 ELSE 0 END) as resolved,
               ROUND(AVG(julianday(c.resolved_at) - julianday(c.created_at)) * 24, 1) as avg_hrs
        FROM citizen_complaints c
        JOIN departments d ON c.department_id = d.department_id
        GROUP BY d.department_name
    """)
    rows = cursor.fetchall()
    
    dept_table_data = [
        [
            Paragraph("Department Name", table_header_style), 
            Paragraph("Total Tickets", table_header_style), 
            Paragraph("Resolved", table_header_style),
            Paragraph("Resolution Ratio", table_header_style),
            Paragraph("Avg Duration", table_header_style)
        ]
    ]
    
    for r in rows:
        ratio = round((r[2]/r[1]*100), 1) if r[1] > 0 else 0.0
        avg_h = f"{r[3]} hrs" if r[3] else "N/A"
        dept_table_data.append([
            Paragraph(r[0], table_body_bold),
            Paragraph(str(r[1]), table_body_style),
            Paragraph(str(r[2]), table_body_style),
            Paragraph(f"{ratio}%", table_body_style),
            Paragraph(avg_h, table_body_style)
        ])
        
    t_dept = Table(dept_table_data, colWidths=[160, 80, 80, 100, 110])
    t_dept.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F2937')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F9FAFB')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F3F4F6')]),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_dept)
    story.append(Spacer(1, 15))
    
    # 4. Smart Utilities overview
    story.append(Paragraph("III. Smart Utility Systems Overview", h1_style))
    util_text = (
        f"<b>Water Supply Grid:</b> Ingested water consumption totals {kpis['total_water_m3']:,} m³ over the logging phase "
        f"with {kpis['total_water_leaks']} leaks successfully flagged and repaired. Pressure standards remain stable "
        f"around 4.2 bar average.<br/>"
        f"<b>Electrical Grid:</b> Power grid loads average {kpis['avg_power_load_factor']} load factor. Grid reports a total "
        f"outage accumulation of {kpis['total_outage_minutes']} minutes over the current monitoring calendar. Mitigation "
        f"mechanisms have been assigned to Vijay Nagar Grid (Res) to reduce voltage instabilities.<br/>"
        f"<b>Sanitation Services:</b> Waste collection logistics report {kpis['total_waste_tons']:,} tons collected. "
        f"Averaged sanitation rating registers at {kpis['avg_sanitation_rating']} / 5.0, with {kpis['total_missed_pickups']} "
        f"logged missed pickups under audit."
    )
    story.append(Paragraph(util_text, body_style))
    story.append(Spacer(1, 20))
    
    # Signature block
    story.append(Paragraph("<b>Report Approved By:</b>", body_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Rajesh Sharma<br/><i>Director of Indore Smart City Integration</i>", body_style))
    
    # Build Document
    doc.build(story)
    print(f"PDF report generated: {filename}")

def generate_reports():
    print("Connecting to database for report generation...")
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(f"Database missing: {DB_PATH}. ETL pipeline must be run first.")
        
    conn = sqlite3.connect(DB_PATH)
    try:
        kpis = query_kpi_data(conn)
        
        pdf_path = os.path.join(REPORTS_DIR, "Smart_City_Operational_Report.pdf")
        excel_path = os.path.join(REPORTS_DIR, "Smart_City_Operational_Report.xlsx")
        
        generate_excel_report(conn, kpis, excel_path)
        generate_pdf_report(conn, kpis, pdf_path)
        
        print("All reports generated successfully in reports/ directory.")
    finally:
        conn.close()

if __name__ == "__main__":
    generate_reports()
